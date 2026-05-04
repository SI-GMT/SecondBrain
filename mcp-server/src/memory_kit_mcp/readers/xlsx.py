"""XLSX reader — extract sheets as Markdown tables via openpyxl.

Mirrors ``scripts/doc-readers/read_xlsx.py``. Each non-empty sheet becomes
a ``## Sheet: {name}`` section. Output is bounded by ``MAX_ROWS`` and
``MAX_COLS`` to keep ``mem_doc`` archives readable; clipping triggers a
warning.
"""

from __future__ import annotations

from pathlib import Path

from . import DocReaderDependencyError

MAX_ROWS = 200
MAX_COLS = 30


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("|", "\\|").strip()


def _render_sheet(name: str, ws, warnings: list[str]) -> str | None:
    rows: list[list[str]] = []
    clipped_rows = clipped_cols = False
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx > MAX_ROWS:
            clipped_rows = True
            break
        cells = [_cell_to_str(c) for c in row]
        if len(cells) > MAX_COLS:
            cells = cells[:MAX_COLS]
            clipped_cols = True
        rows.append(cells)
    while rows and not any(c for c in rows[-1]):
        rows.pop()
    if not rows:
        return None
    if clipped_rows:
        warnings.append(f"sheet {name!r} clipped to {MAX_ROWS} rows")
    if clipped_cols:
        warnings.append(f"sheet {name!r} clipped to {MAX_COLS} columns")
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    out = [f"## Sheet: {name}", ""]
    out.append("| " + " | ".join(rows[0]) + " |")
    out.append("|" + "|".join(["---"] * width) + "|")
    for r in rows[1:]:
        out.append("| " + " | ".join(r) + " |")
    return "\n".join(out)


def extract(path: Path) -> tuple[str, list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise DocReaderDependencyError(
            "openpyxl is required to read .xlsx files. "
            "Install via: pip install memory-kit-mcp[doc-readers]"
        ) from exc

    warnings: list[str] = []
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sections: list[str] = []
        for name in wb.sheetnames:
            ws = wb[name]
            rendered = _render_sheet(name, ws, warnings)
            if rendered:
                sections.append(rendered)
    finally:
        wb.close()
    content = "\n\n".join(sections).strip()
    if not content:
        warnings.append("no content extracted")
    return content, warnings
