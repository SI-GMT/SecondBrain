#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "openpyxl>=3.1",
# ]
# ///
"""
read_xlsx.py — extract content from a .xlsx workbook as Markdown.

Each non-empty sheet becomes a `## Sheet: {name}` section followed by a
Markdown table. The first non-empty row is treated as header.

To keep output bounded, sheets are clipped to MAX_ROWS rows and MAX_COLS
columns (with a stderr warning when clipping occurs). This is a deliberate
trade-off: large data dumps are not the use case of /mem-doc.

Stdout: Markdown.
Stderr: clipping warnings and errors.
Exit codes:
  0  success
  1  invocation error
  2  empty workbook

Convention: invoked by core/procedures/mem-doc.md via `uv run`.
    uv run scripts/doc-readers/read_xlsx.py {path}
"""

import argparse
import sys
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not available. Run via `uv run`.", file=sys.stderr)
    sys.exit(1)


MAX_ROWS = 200
MAX_COLS = 30


def cell_to_str(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("|", "\\|").strip()


def render_sheet(name: str, ws) -> str | None:
    rows: list[list[str]] = []
    clipped_rows = clipped_cols = False
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx > MAX_ROWS:
            clipped_rows = True
            break
        cells = [cell_to_str(c) for c in row]
        if len(cells) > MAX_COLS:
            cells = cells[:MAX_COLS]
            clipped_cols = True
        rows.append(cells)
    while rows and not any(c for c in rows[-1]):
        rows.pop()
    if not rows:
        return None
    if clipped_rows:
        print(f"Warning: sheet '{name}' clipped to {MAX_ROWS} rows", file=sys.stderr)
    if clipped_cols:
        print(f"Warning: sheet '{name}' clipped to {MAX_COLS} columns", file=sys.stderr)
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    out: list[str] = [f"## Sheet: {name}", ""]
    out.append("| " + " | ".join(rows[0]) + " |")
    out.append("|" + "|".join(["---"] * width) + "|")
    for r in rows[1:]:
        out.append("| " + " | ".join(r) + " |")
    return "\n".join(out)


def extract(path: Path) -> str:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    sections: list[str] = []
    for name in wb.sheetnames:
        ws = wb[name]
        rendered = render_sheet(name, ws)
        if rendered:
            sections.append(rendered)
    wb.close()
    return "\n\n".join(sections).strip()


def main() -> int:
    parser = argparse.ArgumentParser(description="Extract .xlsx as Markdown")
    parser.add_argument("path", help="Path to .xlsx file")
    args = parser.parse_args()

    path = Path(args.path)
    if not path.exists():
        print(f"Error: file not found: {path}", file=sys.stderr)
        return 1
    if not path.is_file():
        print(f"Error: not a file: {path}", file=sys.stderr)
        return 1

    try:
        content = extract(path)
    except Exception as exc:
        print(f"Error parsing xlsx: {exc}", file=sys.stderr)
        return 1

    if not content:
        print("Warning: no content extracted", file=sys.stderr)
        return 2

    print(content)
    return 0


if __name__ == "__main__":
    sys.exit(main())
